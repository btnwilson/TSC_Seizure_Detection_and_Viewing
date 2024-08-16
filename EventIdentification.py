
import os 
import pyedflib as edf 
import numpy as np 
import matplotlib.pyplot as plt
from datetime import datetime, timedelta
import pandas as pd
from scipy import signal
from scipy.stats import mode
import time as t
import openpyxl
from openpyxl.drawing.image import Image
import gc
from file_groupings import mice_files
import pickle
from scipy.signal import butter, filtfilt, iirnotch

start_time = t.time()
temp_time1 = t.time()
def calculate_arclength_signal(signal, fs):
    arclength = np.sqrt(np.diff(signal)**2 + (1/fs)**2)
    return arclength

def normalize_median_AD(signal):
    signal_mode, mode_count = mode(signal)
    if mode_count > .05 * len(signal) and signal_mode < .01:
        filtered_signal = signal[(signal != signal_mode)]
    else:
        filtered_signal = signal
    median = np.median(filtered_signal)
    #print(median)
    mad = np.median(np.abs(filtered_signal - median))
    #print(mad)
    normalized_signal = (signal - median)/mad
    return normalized_signal


min_event_duration = 10
max_event_duration = 500
min_spacing = 20
selected_mouse = "cmvcre3_136"
files = mice_files

selected_files = files[selected_mouse]
path = f"Events/{selected_mouse}"
    
cage_number = selected_mouse.split("_")[1]

if not os.path.exists(path):
    os.makedirs(path, exist_ok=True)
    print("Directory Made")
    
for i in range(len(selected_files)):
    #print(i)
    file = selected_files[i]
    print(file)
    signals, headers, main_header = edf.highlevel.read_edf(f"EDF Files/{file}")
    channel1 = signals[0,:]
    fs = headers[0]["sample_frequency"]
    
    normalized_channel1 = normalize_median_AD(channel1)
    
    cutoff = 2
    order = 3
    nyquist = 0.5 * fs
    normal_cutoff = cutoff / nyquist
    b, a = butter(order, normal_cutoff, btype='high', analog=False)
    normalized_channel1 = filtfilt(b, a, normalized_channel1)
    
    f0 = 60
    q = 30
    b, a = iirnotch(f0 / (fs / 2), q, fs)
    normalized_channel1 = filtfilt(b, a, normalized_channel1)
    
    print("File Loaded, Normalized and Filtered")
    
    arclength_signal = calculate_arclength_signal(normalized_channel1, fs)
    time = np.arange(0, len(normalized_channel1)/fs, 1/fs)
    
    mode_normalized, mode_count = mode(normalized_channel1)
    
    if mode_count > .05 * len(normalized_channel1):
        mean_sig = np.mean(normalized_channel1[(normalized_channel1 != mode_normalized)])
        std_sig = np.std(normalized_channel1[(normalized_channel1 != mode_normalized)])
    else:
        mean_sig = np.mean(normalized_channel1)
        std_sig = np.std(normalized_channel1)
        
    threshold = mean_sig + 3 * std_sig
    
    max_peak_inds, max_peak_info = signal.find_peaks(normalized_channel1, height= threshold, distance=50, width=[0, int(.25 * fs)])
    min_peak_inds, min_peak_info = signal.find_peaks(-normalized_channel1, height= threshold, distance=50, width=[0, int(.25 * fs)])
     
    max_peak_heights = max_peak_info["peak_heights"] #["prominences"]
    min_peak_heights = min_peak_info["peak_heights"] #["prominences"]
     
    pos_percentile = np.percentile(max_peak_heights, 75)
    neg_percentile = np.percentile(min_peak_heights, 75)
    
    selected_pos = max_peak_inds[(max_peak_heights >= pos_percentile)] #& (max_peak_heights < pos_95_percentile)]
    selected_neg = min_peak_inds[(min_peak_heights >= neg_percentile)] # & (min_peak_heights < neg_95_percentile)]
    
    time_per_step = 5
    index_step = int(5 * fs)
    
    arclength_sliding_window = np.zeros((int(len(normalized_channel1)/index_step + 1)))
    spikes_sliding_window = np.zeros((int(len(normalized_channel1)/index_step + 1)))
    
    for window_num, index in enumerate(range(0, len(channel1), index_step)):
        arclength_sliding_window[window_num] = np.sum(arclength_signal[index:index+index_step])
        
        total_pos_spikes = np.sum((selected_pos >= index) & (selected_pos < index+index_step)) 
        total_neg_spikes = np.sum((selected_neg >= index) & (selected_neg < index+index_step))
        spikes_sliding_window[window_num] = total_pos_spikes + total_pos_spikes
    
    print("Arclength and Spikes Computed")
    
    spike_events = []
    spike_event = np.array([])
    spike_threshold = 4
    for index, num_spikes in enumerate(spikes_sliding_window):
        index_sec = index * time_per_step
        if num_spikes >= spike_threshold:
            spike_event = np.append(spike_event, index_sec)
        else:
            if len(spike_event) > 0:
                spike_events.append(spike_event)
                spike_event = np.array([])
        if index == len(spikes_sliding_window) - 1 and len(spike_event) > 0:
            spike_events.append(spike_event)
            
    if len(spike_events) == 0:
        print(f"No events found in file:{file}")
        continue
    
    spike_events_merged = []
    current_array = spike_events[0]
    for next_array in spike_events[1:]:
        if next_array[0] - current_array[-1] < min_spacing:
            current_array = np.concatenate((current_array, next_array))
        else:
            spike_events_merged.append(current_array)
            current_array = next_array
    
    spike_events_merged = [event for event in spike_events_merged if event[-1] - event[0] > min_event_duration]
    spike_events_merged = [event for event in spike_events_merged if event[-1] - event[0] < max_event_duration]
    
    max_spikes = []
    max_arclengths = []
    
    for event in spike_events_merged:
        max_spikes.append(max(spikes_sliding_window[int(event[0]/5):int(event[-1]/5)]))
        max_arclengths.append(max(arclength_sliding_window[int(event[0]/5):int(event[-1]/5)]))
    
    max_spikes = np.array(max_spikes)
    max_arclengths = np.array(max_arclengths)
    
    normalized_max_spikes = (max_spikes - np.min(max_spikes))/(np.max(max_spikes) - np.min(max_spikes))
    normalized_max_arclengths = (max_arclengths - np.min(max_arclengths))/(np.max(max_arclengths) - np.min(max_arclengths)) 
    
    weighted_score = normalized_max_spikes + normalized_max_arclengths
    
    ordering = np.argsort(-weighted_score)
    
    spike_events_merged_reordered = [spike_events_merged[j] for j in ordering]
    max_spikes = max_spikes[ordering]
    max_arclengths = max_arclengths[ordering]
    
    file_start_time = main_header["startdate"]
    event_times = []
    
    for event in spike_events_merged_reordered:
        delta_time = timedelta(seconds=event[0])
        time_stamp = file_start_time + delta_time
        event_times.append(time_stamp.strftime("%m-%d-%Y %H:%M:%S"))
        

    if "Done" in file:
        file = [file.removesuffix("Done.edf")] +  [None] * (len(ordering) - 1)
    else:
        file = [file.removesuffix(".edf")] +  [None] * (len(ordering) - 1)


    event_durations = []
    for event in spike_events_merged_reordered:
        event_durations.append(event[-1] - event[0])
        
    output = {"File": file ,
              "Time" : event_times,
              "Max ArcLength" : max_arclengths,
              "Max Spikes" : max_spikes,
              "Event Duration (s)" : event_durations}

    df = pd.DataFrame(output)
    
    print("Data Frame Saved")
    
    #df_short = df.iloc[0:int(len(df) * .2)]
    
    time = np.arange(0, len(channel1)/fs, 1/fs)
    plt.figure()
    plt.plot(time, normalized_channel1)
    
    upper_marker = np.zeros(len(channel1))
    lower_marker = np.zeros(len(channel1))
    #%%
    plot_buffer = 5 * fs
    plt.figure(figsize=(18,12))
    num_plots = 5
    selected_events = {}
    for count, event in enumerate(spike_events_merged_reordered):
        selected_events[f"Event {count}"] = {}
        if event[0] == 0:
            start = 5
        else:
            start = event[0]
        
        if event[-1] >= int(time[-1] - 5):
            end = event[-2]
        else:
            end = event[-1]
            
        event_duration = end - start
        signal_in_window = normalized_channel1[int(start * fs - plot_buffer): int(end * fs + plot_buffer)]
        selected_events[f"Event {count}"]["Start Time"] = event_times[count]
        selected_events[f"Event {count}"]["EEG Signal"] = signal_in_window
        selected_events[f"Event {count}"]["Max Spikes"] = max_spikes[count]
        selected_events[f"Event {count}"]["Max Arclength"] = max_arclengths[count]
        selected_events[f"Event {count}"]["Duration"] = event_durations[count]
        selected_events[f"Event {count}"]["Cage Number"] = cage_number
        
        if count < num_plots:
            plt.subplot(num_plots,1,count+1)
            
            upper_marker = np.zeros(int(event_duration * fs + 10 * fs))
            lower_marker = np.zeros(int(event_duration * fs + 10 * fs))
            
            plot_time = np.arange(0, int(event_duration + 10), 1/fs)
            
            is_event = ((plot_time >= 5) & (plot_time <= event_durations[count] + 5))
            upper_marker[is_event] = max(normalized_channel1[int(start * fs): int(end * fs)])
            lower_marker[is_event] = min(normalized_channel1[int(start * fs): int(end * fs)])
            
            pos_in_window = selected_pos[((selected_pos >= (start * fs - plot_buffer)) & (selected_pos <= (end * fs + plot_buffer)))] - (start * fs) + plot_buffer
            neg_in_window = selected_neg[((selected_neg >= (start * fs - plot_buffer)) & (selected_neg <= (end * fs + plot_buffer)))] - (start * fs) + plot_buffer
            
            plt.plot(plot_time, signal_in_window, label="Normalized Signal")
            plt.fill_between(plot_time, upper_marker, lower_marker, alpha = .5, color="red", label="Identified Event")
            plt.scatter(plot_time[pos_in_window.astype(int)], signal_in_window[pos_in_window.astype(int)], c="red", label="Identified Spike")
            plt.scatter(plot_time[neg_in_window.astype(int)], signal_in_window[neg_in_window.astype(int)], c="red")
            plt.legend()
            plt.xlabel("Time (s)")
            plt.ylabel("Signal Normalized by MAD")
            plt.title(f"Candidate Seizure #{count+1}")
    
    
    plt.tight_layout()
    plt.savefig(f"{file[0]}.png")
    with open(f'Events/{selected_mouse}/{file[0]}.pkl', 'wb') as f:
        pickle.dump(selected_events, f)
        
    excel_path = f"{selected_mouse}.xlsx"
    
    if not os.path.exists(excel_path):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        wb.save(excel_path)
        wb.close()
    
    wb = openpyxl.load_workbook(excel_path)
    
    with pd.ExcelWriter(excel_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        df.to_excel(writer, sheet_name=file[0], index=False)
        
    wb = openpyxl.load_workbook(excel_path)
    new_sheet = wb[f"{file[0]}"]
    
    img = Image(f"{file[0]}.png")
    
    new_sheet.add_image(img, 'H2')
    
    wb.save(excel_path)
    wb.close()
    
    os.remove(f"{file[0]}.png")
    
    
    print(f"{file[0]} Complete")
    temp_time = t.time() - temp_time1
    temp_time1 = t.time()
    print(temp_time)
    
    del channel1
    del signals 
    del normalized_channel1
    del time
    
    gc.collect()
    
end_time = t.time()

elapsed_time = end_time - start_time

print(f"Time taken to run the code: {elapsed_time:.2f} seconds")



